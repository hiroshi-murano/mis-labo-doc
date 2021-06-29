import datetime
import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import PatternFill
from openpyxl.styles import Font


def writeData(cell, txt, color):
    """
    フォント・背景色、罫線を指定してテキストを書き込む
    """

    side = Side(style='thin', color='000000')
    border = Border(top=side, bottom=side, left=side, right=side)
    fill = PatternFill(patternType='solid', fgColor=color)
    font = Font(name='Meiryo UI')

    cell.value = txt
    cell.border = border
    cell.fill = fill
    cell.font = font


def writeText(cell, txt):
    """
    フォントを指定してテキストを書き込む
    """

    font = Font(name='Meiryo UI')
    cell.value = txt
    cell.font = font


def func1():
    """
    """

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'サンプルデータ'

    writeData(ws.cell(row=2, column=1), '伊集院', 'ddFFFF')
    writeData(ws.cell(row=2, column=2), 43, 'FFFFee')
    writeText(ws.cell(row=2, column=3), 54.3)
    writeText(ws.cell(row=2, column=4), datetime.datetime.today())

    wb.save('test4.xlsx')


if __name__ == '__main__':

    func1()
