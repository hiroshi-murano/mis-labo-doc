import openpyxl
import datetime

wb = openpyxl.load_workbook('test.xlsx')
ws = wb['Sheet1']

for row in range(1,  ws.max_row + 1):
    for col in range(1,  ws.max_column + 1):
        data = ws.cell(row=row, column=col).value
        print(data)

wb.close()


wb = openpyxl.load_workbook('test.xlsx')
ws = wb['Sheet1']

for row in ws.iter_rows():
    print(row[0].value)
    print(row[1].value)
    print(row[2].value)
    print(row[3].value)

wb.close()


wb = openpyxl.load_workbook('test.xlsx')
ws = wb['Sheet1']

for row in ws.iter_rows():
    print(row[0].value)
    print(row[1].value)
    print(row[2].value)
    print(row[3].value)

wb.close()


wb = openpyxl.load_workbook('test.xlsx')
ws = wb['Sheet1']

for row in ws.values:
    print(row)

wb.close()


# 既存ファイルへの書き込み
wb = openpyxl.load_workbook('test.xlsx')
ws = wb['Sheet1']

ws.cell(row=2, column=1).value = '伊集院'
ws.cell(row=2, column=2).value = 43
ws.cell(row=2, column=3).value = 59.4
ws.cell(row=2, column=4).value = datetime.datetime.today()

wb.save('test2.xlsx')


# 新規ファイル作成
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'サンプルデータ'

ws.cell(row=2, column=1).value = '伊集院'
ws.cell(row=2, column=2).value = 43
ws.cell(row=2, column=3).value = 59.4
ws.cell(row=2, column=4).value = datetime.datetime.today()

wb.save('test3.xlsx')
